import csv
import io
from datetime import datetime

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.dependencies import get_current_teacher
from app.models.class_student import ClassStudent
from app.models.classroom import Class
from app.models.exam import Exam
from app.models.exam_result import ExamResult
from app.models.student import Student
from app.schemas.student import StudentCreate
from app.services.audit_service import AuditService
from app.services.student_import_service import create_imported_student

router = APIRouter(tags=["Import and export"])


def _exam_and_rows(db: Session, exam_id: int, teacher_id: int):
    exam = db.query(Exam).filter(Exam.id == exam_id, Exam.created_by == teacher_id).first()
    if not exam:
        raise HTTPException(status_code=404, detail="Không tìm thấy đề thi")
    rows = (
        db.query(ExamResult, Student)
        .join(Student, Student.id == ExamResult.student_id)
        .filter(ExamResult.exam_id == exam_id)
        .order_by(Student.student_code, Student.full_name)
        .all()
    )
    return exam, rows


def _table(rows):
    return [
        {
            "student_code": student.student_code or "",
            "student_name": student.full_name or "",
            "email": student.email or "",
            "score": float(result.total_score or 0),
            "started_at": result.started_at,
            "finished_at": result.finished_at,
        }
        for result, student in rows
    ]


@router.get("/results/exam/{exam_id}/export/{format_name}")
def export_results(
    exam_id: int,
    format_name: str,
    teacher=Depends(get_current_teacher),
    db: Session = Depends(get_db),
):
    exam, result_rows = _exam_and_rows(db, exam_id, teacher.id)
    data = _table(result_rows)
    filename = f"ket-qua-{exam_id}"

    if format_name == "csv":
        output = io.StringIO()
        output.write("\ufeff")
        writer = csv.writer(output)
        writer.writerow(["Mã sinh viên", "Họ tên", "Email", "Điểm", "Bắt đầu", "Nộp bài"])
        for row in data:
            writer.writerow([
                row["student_code"], row["student_name"], row["email"], row["score"],
                row["started_at"] or "", row["finished_at"] or "",
            ])
        payload = io.BytesIO(output.getvalue().encode("utf-8"))
        media_type, extension = "text/csv; charset=utf-8", "csv"
    elif format_name == "xlsx":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError as exc:
            raise HTTPException(status_code=500, detail="Chưa cài openpyxl") from exc
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Kết quả"
        sheet.append([exam.title])
        sheet.append(["Mã sinh viên", "Họ tên", "Email", "Điểm", "Bắt đầu", "Nộp bài"])
        for cell in sheet[2]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2563EB")
        for row in data:
            sheet.append([
                row["student_code"], row["student_name"], row["email"], row["score"],
                row["started_at"], row["finished_at"],
            ])
        for column in ("A", "B", "C", "D", "E", "F"):
            sheet.column_dimensions[column].width = 22
        payload = io.BytesIO()
        workbook.save(payload)
        payload.seek(0)
        media_type, extension = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"
        )
    elif format_name == "pdf":
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
        except ImportError as exc:
            raise HTTPException(status_code=500, detail="Chưa cài reportlab") from exc
        payload = io.BytesIO()
        document = SimpleDocTemplate(payload, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        table_data = [["Student ID", "Student name", "Email", "Score", "Submitted"]]
        for row in data:
            table_data.append([
                row["student_code"], row["student_name"], row["email"],
                f'{row["score"]:.2f}', str(row["finished_at"] or ""),
            ])
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        document.build([
            Paragraph(f"Exam results: {exam.title}", styles["Title"]),
            Spacer(1, 12),
            table,
        ])
        payload.seek(0)
        media_type, extension = "application/pdf", "pdf"
    else:
        raise HTTPException(status_code=400, detail="Định dạng phải là csv, xlsx hoặc pdf")

    AuditService.log(db, teacher, "export_results", "exam", exam_id, {"format": format_name})
    return StreamingResponse(
        payload,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}.{extension}"'},
    )


def _read_import(content: bytes, filename: str):
    if filename.lower().endswith(".csv"):
        text = content.decode("utf-8-sig")
        return list(csv.DictReader(io.StringIO(text)))
    if filename.lower().endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise HTTPException(status_code=500, detail="Chưa cài openpyxl") from exc
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        values = list(workbook.active.values)
        if not values:
            return []
        headers = [str(value or "").strip() for value in values[0]]
        return [dict(zip(headers, row)) for row in values[1:]]
    raise HTTPException(status_code=400, detail="Chỉ hỗ trợ tệp CSV hoặc XLSX")


@router.post("/classes/{class_id}/students/import")
async def import_students(
    class_id: int,
    file: UploadFile = File(...),
    teacher=Depends(get_current_teacher),
    db: Session = Depends(get_db),
):
    classroom = db.query(Class).filter(
        Class.id == class_id, Class.teacher_id == teacher.id
    ).first()
    if not classroom:
        raise HTTPException(status_code=404, detail="Không tìm thấy lớp")
    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Tệp vượt quá 5 MB")
    rows = _read_import(content, file.filename or "")
    created, added, errors = 0, 0, []

    for index, row in enumerate(rows, start=2):
        normalized = {
            str(key or "").strip().lower().replace(" ", "_"): value
            for key, value in row.items()
        }
        full_name = normalized.get("full_name") or normalized.get("họ_tên") or normalized.get("ho_ten")
        email = normalized.get("email")
        student_code = (
            normalized.get("student_code") or normalized.get("mã_sinh_viên")
            or normalized.get("ma_sinh_vien")
        )
        try:
            student = None
            if student_code:
                student = db.query(Student).filter(
                    Student.student_code == str(student_code).strip()
                ).first()
            if not student and email:
                student = db.query(Student).filter(
                    Student.email == str(email).strip().lower()
                ).first()
            if not student:
                if not full_name:
                    raise ValueError("thiếu full_name/họ_tên")
                student = create_imported_student(
                    db, str(full_name).strip(), email=email, student_code=student_code
                )
                created += 1
            exists = db.query(ClassStudent).filter(
                ClassStudent.class_id == class_id,
                ClassStudent.student_id == student.id,
            ).first()
            if not exists:
                db.add(ClassStudent(class_id=class_id, student_id=student.id))
                db.commit()
                added += 1
        except Exception as exc:
            db.rollback()
            errors.append({"row": index, "error": str(exc)})

    AuditService.log(
        db, teacher, "import_students", "class", class_id,
        {"created": created, "added": added, "errors": len(errors)},
    )
    return {
        "total": len(rows), "created": created, "added": added,
        "failed": len(errors), "errors": errors[:50],
    }
